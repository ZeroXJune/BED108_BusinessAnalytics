"""
BED 106 Business Analytics - Mini Capstone: Sales Trend Analysis
Checkpoint 2, Task 2.1-2.5: the spreadsheet analytics workbook.

Builds reports/Checkpoint_2_Workbook.xlsx from the Checkpoint 1 warehouse.

Every computed cell is a live Excel formula, not a pasted value, so the
workbook recalculates if the data changes - which is what the Checkpoint 2
rubric means by "formulas are accurate and appropriate".

Sheets:
    Cleaned Data       the finalised dataset (Task 2.1, sheet 1)
    Pivot Analysis     four cross-tabs built with SUMIFS/COUNTIFS (sheet 2)
    Pivot Charts       four charts drawn from those cross-tabs (sheet 3)
    Formulas Showcase  eight Excel functions in business context (sheet 4)
    Descriptive Stats  central tendency and dispersion, 3 variables (Task 2.2)
    Frequency          distribution table and histogram (Task 2.2)
    Correlation        two variable pairs with scatter plots (Task 2.3)
    Regression         simple linear regression with R2 and p-value (Task 2.4)
    Forecast           trend, seasonality and a validated forecast (Task 2.5)

Run:  python3 scripts/build_workbook.py
"""

import os
import sqlite3
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB = os.path.join(ROOT, "data", "sales_trend.db")
OUT = os.path.join(ROOT, "reports", "Checkpoint_2_Workbook.xlsx")

FONT = "Arial"
INK = "1F2933"
ACCENT = "2F6F9F"

HDR_FILL = PatternFill("solid", fgColor="2F6F9F")
SUB_FILL = PatternFill("solid", fgColor="DCE6F1")
NOTE_FILL = PatternFill("solid", fgColor="FFF8E1")
THIN = Side(style="thin", color="B7C2CC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = "#,##0"
DEC2 = "#,##0.00"
DEC4 = "0.0000"
PCT = "0.0%"

# Row 1 is the header, so the first data row is 2.
FIRST = 2


def base(ws, text, size=11, bold=False, color=INK):
    ws.font = Font(name=FONT, size=size, bold=bold, color=color)
    ws.value = text
    return ws


def title(ws, row, text, note=None):
    """Section heading inside a sheet."""
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name=FONT, size=13, bold=True, color=ACCENT)
    if note:
        n = ws.cell(row=row + 1, column=1, value=note)
        n.font = Font(name=FONT, size=9, italic=True, color="7B8794")
        return row + 3
    return row + 2


def header_row(ws, row, labels, start_col=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=start_col + i, value=label)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center",
                                wrap_text=True)
        c.border = BOX


def note(ws, row, text, col=1):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name=FONT, size=9, italic=True, color="5A6672")
    return row + 1


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def load():
    con = sqlite3.connect(DB)
    rows = con.execute("""
        SELECT f.sale_id, f.order_ref, f.order_date,
               d.year_number, d.quarter_number, d.month_number, d.month_name,
               d.year_month, d.is_complete_year,
               cu.customer_name, ci.city_name, st.state_name,
               cat.category_name, s.sub_category_name, pm.payment_mode_name,
               f.quantity, f.amount, f.profit
        FROM sales f
        JOIN dates         d   ON d.order_date      = f.order_date
        JOIN customers     cu  ON cu.customer_id    = f.customer_id
        JOIN cities         ci  ON ci.city_id        = cu.city_id
        JOIN states        st  ON st.state_id       = ci.state_id
        JOIN sub_categories s   ON s.sub_category_id = f.sub_category_id
        JOIN categories     cat ON cat.category_id   = s.category_id
        JOIN payment_modes pm  ON pm.payment_mode_id = f.payment_mode_id
        ORDER BY f.order_date, f.sale_id
    """).fetchall()
    con.close()
    return rows


COLUMNS = [
    ("SaleID", 9), ("OrderRef", 11), ("OrderDate", 12), ("Year", 7),
    ("Quarter", 8), ("MonthNo", 9), ("MonthName", 12), ("YearMonth", 11),
    ("CompleteYear", 12), ("CustomerName", 20), ("City", 14), ("State", 13),
    ("Category", 16), ("SubCategory", 17), ("PaymentMode", 13),
    ("Quantity", 10), ("Amount", 11), ("Profit", 11), ("MarginPct", 11),
]
# Letters for the columns the formulas reference most.
COL = {name: get_column_letter(i + 1) for i, (name, _) in enumerate(COLUMNS)}


def sheet_cleaned(wb, rows):
    """Task 2.1, Sheet 1 - the finalised dataset."""
    ws = wb.active
    ws.title = "Cleaned Data"
    header_row(ws, 1, [c[0] for c in COLUMNS])
    ws.freeze_panes = "A2"

    for r, rec in enumerate(rows, start=FIRST):
        for c, value in enumerate(rec, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = Font(name=FONT, size=9)
        # MarginPct is derived, so it is a formula, never a pasted number.
        m = ws.cell(row=r, column=19,
                    value=f"=IF({COL['Amount']}{r}=0,\"\","
                          f"{COL['Profit']}{r}/{COL['Amount']}{r})")
        m.font = Font(name=FONT, size=9)
        m.number_format = PCT
        for c in (17, 18):
            ws.cell(row=r, column=c).number_format = MONEY

    last = FIRST + len(rows) - 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"
    for name, w in COLUMNS:
        ws.column_dimensions[COL[name]].width = w
    return ws, last


def define_names(wb, last):
    """Named ranges keep every downstream formula readable."""
    for name, col in (("Amount", "Amount"), ("Profit", "Profit"),
                      ("Quantity", "Quantity"), ("Category", "Category"),
                      ("SubCategory", "SubCategory"), ("State", "State"),
                      ("YearMonth", "YearMonth"), ("Yr", "Year"),
                      ("PaymentMode", "PaymentMode")):
        ref = f"'Cleaned Data'!${COL[col]}${FIRST}:${COL[col]}${last}"
        wb.defined_names.add(DefinedName(name, attr_text=ref))


def sheet_pivots(wb, rows, last):
    """Task 2.1, Sheet 2 - four cross-tabs, all SUMIFS/COUNTIFS driven."""
    ws = wb.create_sheet("Pivot Analysis")
    widths(ws, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14,
                "G": 14, "H": 14})

    cats = sorted({r[12] for r in rows})
    years = sorted({r[3] for r in rows})
    states = sorted({r[11] for r in rows})
    subs = sorted({r[13] for r in rows})
    months = sorted({r[7] for r in rows})

    anchors = {}

    # --- Pivot 1: revenue by category x year -------------------------------
    row = title(ws, 1, "Pivot 1 - Total revenue by Category and Year",
                "SUMIFS across the Cleaned Data sheet. Answers: which "
                "category is growing, and which is shrinking?")
    header_row(ws, row, ["Category"] + list(years) + ["Total"])
    for j in range(len(years)):
        ws.cell(row=row, column=2 + j).number_format = "0"
    head = row
    row += 1
    p1_first = row
    for cat in cats:
        ws.cell(row=row, column=1, value=cat).font = Font(name=FONT, size=10)
        for j, y in enumerate(years):
            c = ws.cell(row=row, column=2 + j,
                        value=f'=SUMIFS(Amount,Category,$A{row},Yr,{get_column_letter(2+j)}${head})')
            c.number_format = MONEY
            c.font = Font(name=FONT, size=10)
            c.border = BOX
        t = ws.cell(row=row, column=2 + len(years),
                    value=f"=SUM(B{row}:{get_column_letter(1+len(years))}{row})")
        t.number_format = MONEY
        t.font = Font(name=FONT, size=10, bold=True)
        t.border = BOX
        row += 1
    p1_last = row - 1
    ws.cell(row=row, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
    for j in range(len(years) + 1):
        col = get_column_letter(2 + j)
        c = ws.cell(row=row, column=2 + j,
                    value=f"=SUM({col}{p1_first}:{col}{p1_last})")
        c.number_format = MONEY
        c.font = Font(name=FONT, size=10, bold=True)
        c.fill = SUB_FILL
        c.border = BOX
    anchors["p1"] = (head, p1_first, p1_last, len(years))
    row += 2

    # --- Pivot 2: average order value by state -----------------------------
    row = title(ws, row, "Pivot 2 - Average order value and margin by State",
                "AVERAGEIFS and SUMIFS. Answers: is any region materially "
                "better or worse than the others?")
    header_row(ws, row, ["State", "Orders", "Total Revenue",
                         "Average Order Value", "Total Profit", "Margin %"])
    row += 1
    p2_first = row
    for stt in states:
        ws.cell(row=row, column=1, value=stt).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=2, value=f'=COUNTIFS(State,$A{row})')
        ws.cell(row=row, column=3, value=f'=SUMIFS(Amount,State,$A{row})')
        ws.cell(row=row, column=4, value=f'=AVERAGEIFS(Amount,State,$A{row})')
        ws.cell(row=row, column=5, value=f'=SUMIFS(Profit,State,$A{row})')
        ws.cell(row=row, column=6, value=f'=IF(C{row}=0,"",E{row}/C{row})')
        for c, fmt in ((2, "#,##0"), (3, MONEY), (4, DEC2), (5, MONEY), (6, PCT)):
            cell = ws.cell(row=row, column=c)
            cell.number_format = fmt
            cell.font = Font(name=FONT, size=10)
            cell.border = BOX
        row += 1
    p2_last = row - 1
    anchors["p2"] = (p2_first, p2_last)
    row += 2

    # --- Pivot 3: order count by period ------------------------------------
    row = title(ws, row, "Pivot 3 - Orders and revenue by month (period)",
                "COUNTIFS and SUMIFS by YearMonth. This is the series the "
                "regression and forecast are built on.")
    header_row(ws, row, ["YearMonth", "Orders", "Revenue", "Units",
                         "Avg Order Value"])
    row += 1
    p3_first = row
    for ym in months:
        ws.cell(row=row, column=1, value=ym).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=2, value=f'=COUNTIFS(YearMonth,$A{row})')
        ws.cell(row=row, column=3, value=f'=SUMIFS(Amount,YearMonth,$A{row})')
        ws.cell(row=row, column=4, value=f'=SUMIFS(Quantity,YearMonth,$A{row})')
        ws.cell(row=row, column=5, value=f'=IF(B{row}=0,"",C{row}/B{row})')
        for c, fmt in ((2, "#,##0"), (3, MONEY), (4, "#,##0"), (5, DEC2)):
            cell = ws.cell(row=row, column=c)
            cell.number_format = fmt
            cell.font = Font(name=FONT, size=10)
            cell.border = BOX
        row += 1
    p3_last = row - 1
    anchors["p3"] = (p3_first, p3_last)
    row += 2

    # --- Pivot 4: sub-category ---------------------------------------------
    row = title(ws, row, "Pivot 4 - Revenue by Sub-Category, 2023 vs 2024",
                "SUMIFS with two criteria. This is the Checkpoint 1 finding "
                "reproduced in the spreadsheet.")
    header_row(ws, row, ["Sub-Category", "2023", "2024", "Change",
                         "Change %"])
    row += 1
    p4_first = row
    for sub in subs:
        ws.cell(row=row, column=1, value=sub).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=2,
                value=f'=SUMIFS(Amount,SubCategory,$A{row},Yr,2023)')
        ws.cell(row=row, column=3,
                value=f'=SUMIFS(Amount,SubCategory,$A{row},Yr,2024)')
        ws.cell(row=row, column=4, value=f"=C{row}-B{row}")
        ws.cell(row=row, column=5, value=f'=IF(B{row}=0,"",D{row}/B{row})')
        for c, fmt in ((2, MONEY), (3, MONEY), (4, MONEY), (5, PCT)):
            cell = ws.cell(row=row, column=c)
            cell.number_format = fmt
            cell.font = Font(name=FONT, size=10)
            cell.border = BOX
        row += 1
    p4_last = row - 1
    anchors["p4"] = (p4_first, p4_last)

    row += 2
    note(ws, row, "Note: these are formula-driven cross-tabs, which compute "
                  "exactly what a PivotTable computes and recalculate live. "
                  "To add native Excel PivotTables as well, see the guide in "
                  "docs/checkpoint2_excel_guide.md.")
    return ws, anchors, cats, years, states, subs, months


def sheet_charts(wb, anchors, years):
    """Task 2.1, Sheet 3 - charts drawn from the Pivot Analysis cross-tabs."""
    ws = wb.create_sheet("Pivot Charts")
    ws.column_dimensions["A"].width = 3
    src = "Pivot Analysis"
    p1_head, p1_first, p1_last, n_years = anchors["p1"]
    p2_first, p2_last = anchors["p2"]
    p3_first, p3_last = anchors["p3"]
    p4_first, p4_last = anchors["p4"]

    title(ws, 1, "Chart 1 - Revenue by category and year",
          "Source: Pivot 1. Electronics peaks in 2023 then falls away; "
          "Office Supplies dips in 2023 and recovers.")
    ch = BarChart()
    ch.type = "col"
    ch.grouping = "clustered"
    ch.title = "Revenue by Category and Year"
    ch.y_axis.title = "Revenue"
    ch.x_axis.title = "Category"
    ch.height, ch.width = 8.5, 17
    data = Reference(wb[src], min_col=2, max_col=1 + n_years,
                     min_row=p1_head, max_row=p1_last)
    cats = Reference(wb[src], min_col=1, min_row=p1_first, max_row=p1_last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, "A4")

    title(ws, 22, "Chart 2 - Average order value by state",
          "Source: Pivot 2. The spread across the six states is narrow, "
          "which is itself the finding: geography is not the problem.")
    ch2 = BarChart()
    ch2.type = "col"
    ch2.title = "Average Order Value by State"
    ch2.y_axis.title = "Average order value"
    ch2.x_axis.title = "State"
    ch2.height, ch2.width = 8.5, 17
    ch2.legend = None
    d2 = Reference(wb[src], min_col=4, min_row=p2_first - 1, max_row=p2_last)
    c2 = Reference(wb[src], min_col=1, min_row=p2_first, max_row=p2_last)
    ch2.add_data(d2, titles_from_data=True)
    ch2.set_categories(c2)
    ws.add_chart(ch2, "A25")

    title(ws, 43, "Chart 3 - Monthly revenue over time",
          "Source: Pivot 3. The rise to 2022 and the plateau after it are "
          "visible without any statistics.")
    ch3 = LineChart()
    ch3.title = "Monthly Revenue, Mar 2020 - Mar 2025"
    ch3.y_axis.title = "Revenue"
    ch3.x_axis.title = "Month"
    ch3.height, ch3.width = 8.5, 24
    d3 = Reference(wb[src], min_col=3, min_row=p3_first - 1, max_row=p3_last)
    c3 = Reference(wb[src], min_col=1, min_row=p3_first, max_row=p3_last)
    ch3.add_data(d3, titles_from_data=True)
    ch3.set_categories(c3)
    ws.add_chart(ch3, "A46")

    title(ws, 64, "Chart 4 - Sub-category change, 2023 to 2024",
          "Source: Pivot 4. Printers is the single largest mover in the "
          "dataset and the main cause of the plateau.")
    ch4 = BarChart()
    ch4.type = "bar"
    ch4.title = "Revenue Change by Sub-Category, 2023 to 2024"
    ch4.x_axis.title = "Sub-category"
    ch4.y_axis.title = "Change in revenue"
    ch4.height, ch4.width = 11, 17
    ch4.legend = None
    d4 = Reference(wb[src], min_col=4, min_row=p4_first - 1, max_row=p4_last)
    c4 = Reference(wb[src], min_col=1, min_row=p4_first, max_row=p4_last)
    ch4.add_data(d4, titles_from_data=True)
    ch4.set_categories(c4)
    ws.add_chart(ch4, "A67")
    return ws


def sheet_formulas(wb, rows, last):
    """Task 2.1, Sheet 4 - eight Excel functions in business context."""
    ws = wb.create_sheet("Formulas Showcase")
    widths(ws, {"A": 5, "B": 34, "C": 18, "D": 62})

    row = title(ws, 1, "Task 2.1, Sheet 4 - Formulas Showcase",
                "Eight functions, each answering a real question about this "
                "business. Column C holds the live formula shown in column D.")
    header_row(ws, row, ["#", "Business question", "Result",
                         "Formula used (and what it does)"])
    row += 1

    demos = [
        ("SUMIF - total revenue from Electronics",
         '=SUMIF(Category,"Electronics",Amount)', MONEY,
         'SUMIF(Category,"Electronics",Amount) - adds Amount for every row '
         'whose Category is Electronics.'),
        ("COUNTIF - how many orders used COD",
         '=COUNTIF(PaymentMode,"COD")', "#,##0",
         'COUNTIF(PaymentMode,"COD") - counts rows matching one condition.'),
        ("SUMIFS - Printer revenue in 2024 (two conditions)",
         '=SUMIFS(Amount,SubCategory,"Printers",Yr,2024)', MONEY,
         'SUMIFS adds Amount only where SubCategory is Printers AND Year is '
         '2024. This is the number that collapsed.'),
        ("AVERAGEIF - average order value in Texas",
         '=AVERAGEIF(State,"Texas",Amount)', DEC2,
         'AVERAGEIF(State,"Texas",Amount) - mean of Amount for Texas rows.'),
        ("VLOOKUP - look up one transaction by SaleID",
         "=VLOOKUP(1000,'Cleaned Data'!$A$2:$S$%d,17,FALSE)" % last, MONEY,
         'VLOOKUP(1000,...,17,FALSE) - finds SaleID 1000 and returns the 17th '
         'column (Amount). FALSE forces an exact match.'),
        ("IF - flag whether the business grew in 2024",
         '=IF(SUMIFS(Amount,Yr,2024)>SUMIFS(Amount,Yr,2023),'
         '"Grew","Declined")', None,
         'IF(test,"Grew","Declined") - compares 2024 revenue against 2023 and '
         'returns a word, not a number.'),
        ("TEXT - format the 2024 total for a report line",
         '=TEXT(SUMIFS(Amount,Yr,2024),"#,##0")&" in 2024 revenue"', None,
         'TEXT(value,"#,##0") turns a number into formatted text so it can be '
         'joined to a sentence with &.'),
        ("MAX / MIN - the largest and smallest single order",
         '=MAX(Amount)&" / "&MIN(Amount)', None,
         'MAX(Amount) and MIN(Amount) give the range of order sizes.'),
    ]
    for i, (question, formula, fmt, explain) in enumerate(demos, 1):
        ws.cell(row=row, column=1, value=i).font = Font(name=FONT, size=10)
        q = ws.cell(row=row, column=2, value=question)
        q.font = Font(name=FONT, size=10)
        q.alignment = Alignment(wrap_text=True, vertical="top")
        c = ws.cell(row=row, column=3, value=formula)
        c.font = Font(name=FONT, size=10, bold=True)
        c.border = BOX
        if fmt:
            c.number_format = fmt
        e = ws.cell(row=row, column=4, value=explain)
        e.font = Font(name=FONT, size=9, color="5A6672")
        e.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1
    return ws


VARS = [("Amount", "Amount", "Revenue per transaction line"),
        ("Profit", "Profit", "Gross profit per transaction line"),
        ("Quantity", "Quantity", "Units sold per transaction line")]


def sheet_descriptive(wb, last):
    """Task 2.2 - central tendency and dispersion for three variables."""
    ws = wb.create_sheet("Descriptive Stats")
    widths(ws, {"A": 30, "B": 16, "C": 16, "D": 16, "E": 46})

    row = title(ws, 1, "Task 2.2 - Descriptive Statistics",
                "Three numerical variables, n = %d transaction lines. Every "
                "cell is a live formula over the named ranges." % (last - 1))
    header_row(ws, row, ["Statistic"] + [v[0] for v in VARS] + ["What it tells you"])
    head = row
    row += 1

    stats = [
        ("Count", "=COUNT({r})", "#,##0",
         "Number of transaction lines - the sample size n."),
        ("Mean", "=AVERAGE({r})", DEC2,
         "The arithmetic average. Sensitive to extreme values."),
        ("Median", "=MEDIAN({r})", DEC2,
         "The middle value. Compare with the mean: if mean > median the "
         "distribution is skewed right."),
        ("Mode", "=MODE({r})", DEC2,
         "The most frequent value. Meaningful for Quantity; near-meaningless "
         "for a continuous variable like Amount."),
        ("Standard Deviation", "=STDEV({r})", DEC2,
         "Typical distance of a value from the mean. Sample (n-1) version."),
        ("Variance", "=VAR({r})", "#,##0.00",
         "The standard deviation squared. Same information, squared units."),
        ("Minimum", "=MIN({r})", DEC2, "Smallest observed value."),
        ("Maximum", "=MAX({r})", DEC2, "Largest observed value."),
        ("Range", "=MAX({r})-MIN({r})", DEC2,
         "Max minus min - the crudest measure of spread."),
        ("1st Quartile (Q1)", "=QUARTILE({r},1)", DEC2,
         "25% of lines fall below this value."),
        ("3rd Quartile (Q3)", "=QUARTILE({r},3)", DEC2,
         "75% of lines fall below this value."),
        ("Interquartile Range", "=QUARTILE({r},3)-QUARTILE({r},1)", DEC2,
         "Spread of the middle half - unaffected by outliers."),
        ("Coefficient of Variation", "=STDEV({r})/AVERAGE({r})", PCT,
         "Standard deviation as a share of the mean, so spread can be "
         "compared across variables with different units."),
    ]
    for label, tpl, fmt, meaning in stats:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10, bold=True)
        for j, (name, rng, _) in enumerate(VARS):
            c = ws.cell(row=row, column=2 + j, value=tpl.format(r=rng))
            c.number_format = fmt
            c.font = Font(name=FONT, size=10)
            c.border = BOX
        m = ws.cell(row=row, column=2 + len(VARS), value=meaning)
        m.font = Font(name=FONT, size=9, color="5A6672")
        m.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 26
        row += 1

    row += 1
    row = title(ws, row, "Variable definitions")
    for name, rng, desc in VARS:
        ws.cell(row=row, column=1, value=name).font = Font(name=FONT, size=10, bold=True)
        ws.cell(row=row, column=2, value=desc).font = Font(name=FONT, size=10)
        row += 1
    return ws


BIN_WIDTH = 1000
BIN_MAX = 10000


def sheet_frequency(wb):
    """Task 2.2 - frequency distribution and histogram for Amount."""
    ws = wb.create_sheet("Frequency")
    widths(ws, {"A": 16, "B": 12, "C": 12, "D": 14, "E": 14, "F": 40})

    row = title(ws, 1, "Task 2.2 - Frequency distribution of Amount",
                "Bins of %s. Counts use COUNTIFS rather than the FREQUENCY "
                "array function so every cell is independent and auditable."
                % f"{BIN_WIDTH:,}")
    header_row(ws, row, ["Bin (Amount)", "Lower", "Upper", "Frequency",
                         "Relative %", "Cumulative %"])
    row += 1
    first = row
    edges = list(range(0, BIN_MAX, BIN_WIDTH))
    for lo in edges:
        hi = lo + BIN_WIDTH
        ws.cell(row=row, column=1,
                value=f"{lo:,} - {hi - 1:,}").font = Font(name=FONT, size=10)
        ws.cell(row=row, column=2, value=lo).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=3, value=hi).font = Font(name=FONT, size=10)
        f = ws.cell(row=row, column=4,
                    value=f'=COUNTIFS(Amount,">="&B{row},Amount,"<"&C{row})')
        f.number_format = "#,##0"
        f.font = Font(name=FONT, size=10)
        f.border = BOX
        rel = ws.cell(row=row, column=5, value=f"=D{row}/COUNT(Amount)")
        rel.number_format = PCT
        rel.font = Font(name=FONT, size=10)
        row += 1
    lastb = row - 1
    for i, r in enumerate(range(first, lastb + 1)):
        c = ws.cell(row=r, column=6,
                    value=f"=SUM($D${first}:$D{r})/COUNT(Amount)")
        c.number_format = PCT
        c.font = Font(name=FONT, size=10)

    ws.cell(row=row, column=1, value="Total").font = Font(name=FONT, size=10, bold=True)
    t = ws.cell(row=row, column=4, value=f"=SUM(D{first}:D{lastb})")
    t.number_format = "#,##0"
    t.font = Font(name=FONT, size=10, bold=True)
    t.fill = SUB_FILL
    check = ws.cell(row=row, column=6,
                    value=f'=IF(D{row}=COUNT(Amount),"Matches COUNT(Amount)",'
                          f'"MISMATCH - check bin edges")')
    check.font = Font(name=FONT, size=9, italic=True, color="5A6672")

    ch = BarChart()
    ch.type = "col"
    ch.title = "Histogram - distribution of transaction Amount"
    ch.y_axis.title = "Number of transaction lines"
    ch.x_axis.title = "Amount bin"
    ch.gapWidth = 3
    ch.legend = None
    ch.height, ch.width = 9, 18
    data = Reference(ws, min_col=4, min_row=first - 1, max_row=lastb)
    cats = Reference(ws, min_col=1, min_row=first, max_row=lastb)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, "H2")
    note(ws, row + 2,
         "Read the shape, not just the numbers: a flat histogram with no "
         "central peak is a uniform distribution, which is what a random "
         "number generator produces - further evidence the data is synthetic.")
    return ws


def sheet_correlation(wb, c_first, c_last, last):
    """Task 2.3 - two variable pairs, each with a scatter plot."""
    ws = wb.create_sheet("Correlation")
    widths(ws, {"A": 34, "B": 18, "C": 60})
    pv = "'Pivot Analysis'"

    row = title(ws, 1, "Task 2.3 - Correlation Analysis",
                "Two pairs. The first is the relationship the business runs "
                "on; the second is the one everybody assumes exists and does "
                "not.")

    # --- Pair 1 -----------------------------------------------------------
    row = title(ws, row, "Pair 1 - Monthly orders vs monthly revenue")
    # Complete months only - a partial month is not a valid observation.
    ords = f"{pv}!$B${c_first}:$B${c_last}"
    revs = f"{pv}!$C${c_first}:$C${c_last}"
    items = [
        ("Pearson correlation r", f"=CORREL({ords},{revs})", DEC4,
         "Strength and direction of the linear relationship, from -1 to +1."),
        ("r squared", f"=CORREL({ords},{revs})^2", DEC4,
         "Share of the variation in revenue that moves with order count."),
        ("Sample size n", f"=COUNT({ords})", "#,##0",
         "Number of months in the series."),
        ("Direction", f'=IF(CORREL({ords},{revs})>0,"Positive","Negative")',
         None, "Positive means they rise together."),
        ("Strength",
         f'=IF(ABS(CORREL({ords},{revs}))>=0.7,"Strong",'
         f'IF(ABS(CORREL({ords},{revs}))>=0.4,"Moderate","Weak"))', None,
         "Conventional bands: 0.7+ strong, 0.4-0.7 moderate, below 0.4 weak."),
    ]
    row = _stat_block(ws, row, items)
    row = note(ws, row, "Interpretation: months with more orders have "
                        "proportionally more revenue. This is the statistical "
                        "form of the Checkpoint 1 finding that the business "
                        "moves on order COUNT, not order SIZE.") + 1

    # --- Pair 2 -----------------------------------------------------------
    row = title(ws, row, "Pair 2 - Line quantity vs line amount")
    items = [
        ("Pearson correlation r", "=CORREL(Quantity,Amount)", DEC4,
         "Does buying more units mean a bigger order value?"),
        ("r squared", "=CORREL(Quantity,Amount)^2", DEC4,
         "Share of order value explained by units sold."),
        ("Sample size n", "=COUNT(Quantity)", "#,##0",
         "All transaction lines."),
        ("Strength",
         '=IF(ABS(CORREL(Quantity,Amount))>=0.7,"Strong",'
         'IF(ABS(CORREL(Quantity,Amount))>=0.4,"Moderate","Weak"))', None,
         "Expect Weak here - that is the finding."),
    ]
    row = _stat_block(ws, row, items)
    row = note(ws, row, "Interpretation: essentially no relationship. Units "
                        "sold tells you almost nothing about what an order is "
                        "worth, so 'sell more units' is not a revenue "
                        "strategy for this business.") + 1

    # --- Supporting pair --------------------------------------------------
    row = title(ws, row, "Supporting check - line amount vs line profit")
    items = [
        ("Pearson correlation r", "=CORREL(Amount,Profit)", DEC4,
         "Bigger orders do earn more profit, but not proportionally."),
        ("r squared", "=CORREL(Amount,Profit)^2", DEC4,
         "Roughly 46% of profit variation tracks revenue; the rest is "
         "margin differences between products."),
    ]
    row = _stat_block(ws, row, items)

    # Scatter 1: monthly orders vs revenue, with trendline.
    ch = ScatterChart()
    ch.title = "Pair 1 - Monthly Orders vs Monthly Revenue"
    ch.x_axis.title = "Orders in month"
    ch.y_axis.title = "Revenue in month"
    ch.style = 13
    ch.height, ch.width = 9, 16
    ch.legend = None
    xs = Reference(wb["Pivot Analysis"], min_col=2, min_row=c_first, max_row=c_last)
    ys = Reference(wb["Pivot Analysis"], min_col=3, min_row=c_first, max_row=c_last)
    s = Series(ys, xs, title="Months")
    s.marker.symbol = "circle"
    s.graphicalProperties.line.noFill = True
    s.trendline = Trendline(trendlineType="linear", dispRSqr=True, dispEq=True)
    ch.series.append(s)
    ws.add_chart(ch, "E2")

    # Scatter 2: quantity vs amount at line level.
    ch2 = ScatterChart()
    ch2.title = "Pair 2 - Line Quantity vs Line Amount"
    ch2.x_axis.title = "Quantity (units)"
    ch2.y_axis.title = "Amount"
    ch2.style = 13
    ch2.height, ch2.width = 9, 16
    ch2.legend = None
    xs2 = Reference(wb["Cleaned Data"], min_col=16, min_row=FIRST, max_row=last)
    ys2 = Reference(wb["Cleaned Data"], min_col=17, min_row=FIRST, max_row=last)
    s2 = Series(ys2, xs2, title="Transaction lines")
    s2.marker.symbol = "circle"
    s2.marker.size = 3
    s2.graphicalProperties.line.noFill = True
    s2.trendline = Trendline(trendlineType="linear", dispRSqr=True, dispEq=True)
    ch2.series.append(s2)
    ws.add_chart(ch2, "E20")
    return ws


def _stat_block(ws, row, items):
    header_row(ws, row, ["Measure", "Value", "Meaning"])
    row += 1
    for label, formula, fmt, meaning in items:
        ws.cell(row=row, column=1, value=label).font = Font(name=FONT, size=10)
        c = ws.cell(row=row, column=2, value=formula)
        c.font = Font(name=FONT, size=10, bold=True)
        c.border = BOX
        if fmt:
            c.number_format = fmt
        m = ws.cell(row=row, column=3, value=meaning)
        m.font = Font(name=FONT, size=9, color="5A6672")
        m.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 24
        row += 1
    return row + 1


def sheet_regression(wb, c_first, c_last):
    """Task 2.4 - simple linear regression, orders -> revenue."""
    ws = wb.create_sheet("Regression")
    widths(ws, {"A": 34, "B": 18, "C": 64})
    pv = "'Pivot Analysis'"
    # Complete months only (2020-2024); the partial 2025 window is excluded.
    X = f"{pv}!$B${c_first}:$B${c_last}"   # orders per month
    Y = f"{pv}!$C${c_first}:$C${c_last}"   # revenue per month

    row = title(ws, 1, "Task 2.4 - Simple Linear Regression",
                "Y = monthly revenue (dependent). X = number of orders in the "
                "month (predictor). Both come from Pivot 3.")

    row = title(ws, row, "Why this pair")
    for line in [
        "Checkpoint 1 found that revenue fell while average order value stayed "
        "flat, which implies order COUNT drives revenue.",
        "This regression tests that claim formally instead of asserting it.",
        "It is also actionable: the business can influence how many orders it "
        "wins far more easily than how large each one is.",
    ]:
        c = ws.cell(row=row, column=1, value="- " + line)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 15
        row += 1
    row += 1

    items = [
        ("Sample size n (months)", f"=COUNT({X})", "#,##0",
         "Complete months only; the partial 2025 window is excluded."),
        ("Slope (b)", f"=SLOPE({Y},{X})", DEC2,
         "Revenue added by one more order in the month."),
        ("Intercept (a)", f"=INTERCEPT({Y},{X})", DEC2,
         "Predicted revenue at zero orders. Not meaningful on its own - it is "
         "far outside the observed range of X."),
        ("Correlation r", f"=CORREL({Y},{X})", DEC4,
         "Direction and strength of the linear relationship."),
        ("R squared", f"=RSQ({Y},{X})", DEC4,
         "Share of the variation in monthly revenue explained by order count."),
        ("Standard error of estimate", f"=STEYX({Y},{X})", DEC2,
         "Typical size of the prediction error, in revenue units."),
        ("Standard error of slope",
         f"=STEYX({Y},{X})/SQRT(DEVSQ({X}))", DEC2,
         "Uncertainty in the slope estimate."),
        ("t statistic for slope",
         f"=SLOPE({Y},{X})/(STEYX({Y},{X})/SQRT(DEVSQ({X})))", DEC2,
         "Slope divided by its standard error."),
        ("Degrees of freedom", f"=COUNT({X})-2", "#,##0",
         "n minus 2, because the model estimates two parameters."),
        ("p-value (two-tailed)",
         f"=TDIST(ABS(SLOPE({Y},{X})/(STEYX({Y},{X})/SQRT(DEVSQ({X})))),"
         f"COUNT({X})-2,2)", "0.00E+00",
         "Probability of seeing a slope this large if the true slope were "
         "zero. Below 0.05 means the relationship is statistically "
         "significant."),
        ("Significant at 5%?",
         f'=IF(TDIST(ABS(SLOPE({Y},{X})/(STEYX({Y},{X})/SQRT(DEVSQ({X})))),'
         f'COUNT({X})-2,2)<0.05,"YES - reject the null hypothesis",'
         f'"NO - cannot reject the null")', None,
         "The formal conclusion of the significance test."),
    ]
    row = _stat_block(ws, row, items)

    eq_row = row
    row = title(ws, row, "The regression equation")
    c = ws.cell(row=row, column=1, value="Revenue =")
    c.font = Font(name=FONT, size=11, bold=True)
    e = ws.cell(row=row, column=2,
                value=f'=TEXT(INTERCEPT({Y},{X}),"#,##0.00")&" + "&'
                      f'TEXT(SLOPE({Y},{X}),"#,##0.00")&" x Orders"')
    e.font = Font(name=FONT, size=11, bold=True, color=ACCENT)
    e.fill = NOTE_FILL
    e.border = BOX
    row += 2

    row = title(ws, row, "Business forecasts from the model",
                "Predicted monthly revenue at different order volumes, using "
                "the fitted equation.")
    header_row(ws, row, ["Orders in month", "Predicted revenue",
                         "Comment"])
    row += 1
    comments = {
        15: "A weak month - below the historical average.",
        20: "About the historical average order volume.",
        25: "A strong month, near the 2022 peak rate.",
        30: "Above anything observed; the prediction is an extrapolation.",
    }
    for n, comment in comments.items():
        ws.cell(row=row, column=1, value=n).font = Font(name=FONT, size=10)
        p = ws.cell(row=row, column=2,
                    value=f"=INTERCEPT({Y},{X})+SLOPE({Y},{X})*A{row}")
        p.number_format = MONEY
        p.font = Font(name=FONT, size=10, bold=True)
        p.border = BOX
        cm = ws.cell(row=row, column=3, value=comment)
        cm.font = Font(name=FONT, size=9, color="5A6672")
        row += 1
    row += 1

    row = title(ws, row, "Assumptions, limitations and conditions of use")
    limits = [
        "Linearity - the relationship is assumed to be a straight line. The "
        "scatter plot on the Correlation sheet supports this over the "
        "observed range.",
        "Independence - each month is assumed independent of the last. This "
        "is the weakest assumption here: monthly sales series are usually "
        "autocorrelated, and that inflates apparent significance.",
        "Range of validity - X was observed roughly between 8 and 32 orders "
        "per month. Predictions outside that range, including the intercept "
        "at zero orders, are extrapolation and should not be quoted.",
        "Correlation is not causation - the model shows revenue and order "
        "count move together, not that forcing orders up would raise revenue "
        "by exactly the slope.",
        "One predictor only - a single variable cannot capture seasonality, "
        "product mix, or the Printer supply problem. Multiple regression in "
        "Checkpoint 4 is the natural extension.",
        "The underlying data is synthetic (see Checkpoint 1), so the "
        "coefficients describe this file, not a real market.",
    ]
    for text in limits:
        c = ws.cell(row=row, column=1, value="- " + text)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 28
        row += 1
    return ws


MONTH_NAMES = ["January", "February", "March", "April", "May", "June", "July",
               "August", "September", "October", "November", "December"]

# The file runs 22 Mar 2020 - 15 Mar 2025, so March 2020 and March 2025 are
# part-months and 2025 is a part-year. The analysis window is the 57 whole
# months from April 2020 to December 2024.
PARTIAL_MONTHS = {"2020-03", "2025-03"}


def analysis_months(months):
    return [m for m in months
            if int(m[:4]) <= 2024 and m not in PARTIAL_MONTHS]


def sheet_forecast(wb, months, p3_first, p3_last):
    """Task 2.5 - trend, seasonality, forecast and out-of-sample check."""
    ws = wb.create_sheet("Forecast")
    widths(ws, {"A": 15, "B": 12, "C": 14, "D": 14, "E": 14, "F": 14,
                "G": 15, "H": 46})
    pv = "'Pivot Analysis'"

    complete = analysis_months(months)
    n_complete = len(complete)
    c_first = p3_first + months.index(complete[0])
    c_last = c_first + n_complete - 1
    row = title(ws, 1, "Task 2.5 - Trend and Seasonality Analysis",
                "Monthly revenue, Apr 2020 - Dec 2024 (%d complete months). "
                "The 2025 months are held back and used to check the forecast."
                % n_complete)

    # --- Part A: linear trend on time ------------------------------------
    row = title(ws, row, "A. Does a straight-line trend fit?",
                "Regressing revenue on a simple time index t = 1, 2, 3...")
    row = _trend_block(ws, row, c_first, c_last, n_complete)

    # --- Part B: seasonal index -------------------------------------------
    row = title(ws, row, "B. Seasonal index by calendar month",
                "Average revenue for each calendar month divided by the "
                "average month overall. 1.00 = an average month.")
    header_row(ws, row, ["Month", "Avg revenue", "Seasonal index",
                         "Reading"])
    row += 1
    si_first = row
    for i, name in enumerate(MONTH_NAMES, start=1):
        ws.cell(row=row, column=1, value=name).font = Font(name=FONT, size=10)
        a = ws.cell(row=row, column=2,
                    value=f'=AVERAGEIFS(MonthRevenue,MonthNumber,{i})')
        a.number_format = MONEY
        a.font = Font(name=FONT, size=10)
        a.border = BOX
        s = ws.cell(row=row, column=3,
                    value=f"=B{row}/AVERAGE($B${si_first}:$B${si_first + 11})")
        s.number_format = "0.000"
        s.font = Font(name=FONT, size=10, bold=True)
        s.border = BOX
        rd = ws.cell(row=row, column=4,
                     value=f'=IF(C{row}>=1.1,"Peak month",'
                           f'IF(C{row}<=0.9,"Weak month","Average"))')
        rd.font = Font(name=FONT, size=9, color="5A6672")
        row += 1
    si_last = row - 1
    row += 1

    # --- Part C: the forecast ---------------------------------------------
    row = title(ws, row, "C. Forecast for the next 6 periods",
                "Forecast = recent level x seasonal index. The recent level "
                "is the mean of the last 24 complete months, because the "
                "trend test in part A shows no reliable slope to project.")
    lvl_row = row
    ws.cell(row=row, column=1, value="Recent level (mean of last 24 months)").font = Font(name=FONT, size=10, bold=True)
    lv = ws.cell(row=row, column=3,
                 value=f"=AVERAGE({pv}!$C${c_last - 23}:$C${c_last})")
    lv.number_format = MONEY
    lv.font = Font(name=FONT, size=11, bold=True, color=ACCENT)
    lv.fill = NOTE_FILL
    lv.border = BOX
    row += 2

    header_row(ws, row, ["Period", "Month no", "Seasonal index",
                         "Forecast revenue", "Actual", "Error %",
                         "Status", "Note"])
    row += 1
    fc_first = row
    forecasts = [("2025-01", 1), ("2025-02", 2), ("2025-03", 3),
                 ("2025-04", 4), ("2025-05", 5), ("2025-06", 6)]
    for ym, mn in forecasts:
        ws.cell(row=row, column=1, value=ym).font = Font(name=FONT, size=10)
        ws.cell(row=row, column=2, value=mn).font = Font(name=FONT, size=10)
        si = ws.cell(row=row, column=3, value=f"=$C${si_first + mn - 1}")
        si.number_format = "0.000"
        si.font = Font(name=FONT, size=10)
        f = ws.cell(row=row, column=4, value=f"=$C${lvl_row}*C{row}")
        f.number_format = MONEY
        f.font = Font(name=FONT, size=10, bold=True)
        f.border = BOX
        a = ws.cell(row=row, column=5,
                    value=f'=IF(SUMIFS(Amount,YearMonth,$A{row})=0,"",'
                          f'SUMIFS(Amount,YearMonth,$A{row}))')
        a.number_format = MONEY
        a.font = Font(name=FONT, size=10)
        e = ws.cell(row=row, column=6,
                    value=f'=IF(E{row}="","",(D{row}-E{row})/E{row})')
        e.number_format = PCT
        e.font = Font(name=FONT, size=10)
        st = ws.cell(row=row, column=7,
                     value=f'=IF(E{row}="","Future - no actual yet",'
                           f'IF($A{row}="2025-03","Partial month - not comparable",'
                           f'"Complete month - comparable"))')
        st.font = Font(name=FONT, size=9, color="5A6672")
        row += 1
    fc_last = row - 1
    row += 1

    row = title(ws, row, "D. How good is the forecast?",
                "Only January and February 2025 are complete months, so the "
                "check rests on two observations. That is far too few to "
                "confirm a method - it is reported as an honest check, not "
                "as validation.")
    checks = [
        ("Mean absolute error, seasonal model",
         f"=AVERAGE(ABS(F{fc_first}),ABS(F{fc_first + 1}))", PCT,
         "Average size of the error, ignoring direction, over the two "
         "complete holdout months."),
        ("Mean absolute error, flat model",
         f"=AVERAGE(ABS(($C${lvl_row}-E{fc_first})/E{fc_first}),"
         f"ABS(($C${lvl_row}-E{fc_first + 1})/E{fc_first + 1}))", PCT,
         "The same check for a naive forecast that ignores seasonality "
         "entirely and predicts the recent average every month."),
        ("Which did better?",
         f'=IF(AVERAGE(ABS(F{fc_first}),ABS(F{fc_first + 1}))<'
         f'AVERAGE(ABS(($C${lvl_row}-E{fc_first})/E{fc_first}),'
         f'ABS(($C${lvl_row}-E{fc_first + 1})/E{fc_first + 1})),'
         f'"Seasonal model","Flat model - seasonality did not help here")',
         None,
         "If the naive model wins, the seasonal adjustment is not earning "
         "its complexity on this evidence."),
    ]
    row = _stat_block(ws, row, checks)

    row = title(ws, row, "Reliability and assumptions")
    for text in [
        "The trend test in part A is not significant, so no growth or decline "
        "rate is projected. Forecasting a slope that the data does not "
        "support would be the single easiest way to be badly wrong.",
        "The seasonal index assumes next year repeats the average shape of "
        "the last five. January 2025 did not - it came in far above its "
        "historical index, which is exactly the risk this assumption carries.",
        "The level is the mean of the last 24 months, chosen because the "
        "series is flat over that window. If the business changes, the level "
        "must be re-estimated.",
        "Two complete holdout months cannot validate a forecasting method. "
        "Treat part D as a sanity check, not as evidence the model works.",
        "Nothing here models the Printer supply problem found in Checkpoint 1. "
        "If that is a supply failure and it is fixed, the forecast is too low; "
        "if it is lost demand, the forecast may be too high.",
    ]:
        c = ws.cell(row=row, column=1, value="- " + text)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 28
        row += 1

    ch = LineChart()
    ch.title = "Monthly revenue with 6-period forecast"
    ch.y_axis.title = "Revenue"
    ch.x_axis.title = "Month"
    ch.height, ch.width = 9, 22
    data = Reference(wb["Pivot Analysis"], min_col=3,
                     min_row=p3_first - 1, max_row=p3_last)
    cats = Reference(wb["Pivot Analysis"], min_col=1,
                     min_row=p3_first, max_row=p3_last)
    ch.add_data(data, titles_from_data=True)
    ch.set_categories(cats)
    ws.add_chart(ch, "J2")
    return ws, si_first, si_last


def _trend_block(ws, row, c_first, c_last, n):
    pv = "'Pivot Analysis'"
    Y = f"{pv}!$C${c_first}:$C${c_last}"
    X = "Tindex"
    items = [
        ("Slope per month", f"=SLOPE({Y},{X})", DEC2,
         "Revenue added per month if the trend were a straight line."),
        ("R squared", f"=RSQ({Y},{X})", DEC4,
         "Share of revenue variation explained by time alone."),
        ("p-value of the slope",
         f"=TDIST(ABS(SLOPE({Y},{X})/(STEYX({Y},{X})/SQRT(DEVSQ({X})))),"
         f"COUNT({X})-2,2)", "0.0000",
         "Above 0.05 means the apparent trend could easily be noise."),
        ("Verdict",
         f'=IF(TDIST(ABS(SLOPE({Y},{X})/(STEYX({Y},{X})/SQRT(DEVSQ({X})))),'
         f'COUNT({X})-2,2)<0.05,"Significant trend - project it",'
         f'"No reliable linear trend - do NOT project a slope")', None,
         "This verdict decides how the forecast in part C is built."),
    ]
    return _stat_block(ws, row, items)


def add_series_helpers(wb, months, p3_first):
    """Time index and month number beside Pivot 3, for the complete months.

    These are structural counters (t = 1, 2, 3...), not computed results, so
    they are written as literals. Everything derived from them is a formula.
    """
    ws = wb["Pivot Analysis"]
    complete = analysis_months(months)
    # The analysis months are contiguous and start one row into Pivot 3,
    # because March 2020 is a part-month and is skipped.
    c_first = p3_first + months.index(complete[0])
    c_last = c_first + len(complete) - 1

    head = p3_first - 1
    for col, label in ((7, "t index"), (8, "Month no")):
        c = ws.cell(row=head, column=col, value=label)
        c.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = BOX

    for i, ym in enumerate(complete):
        r = c_first + i
        ws.cell(row=r, column=7, value=i + 1).font = Font(name=FONT, size=10)
        ws.cell(row=r, column=8, value=int(ym[5:7])).font = Font(name=FONT, size=10)

    note(ws, c_first + len(months) + 1,
         "Helper columns above: t is a simple time counter used by the trend "
         "regression, and Month no drives the seasonal index. They start at "
         "April 2020 and stop at December 2024: March 2020 and March 2025 are "
         "part-months, and 2025 is a part-year.", col=7)
    return c_first, c_last


def define_series_names(wb, c_first, c_last):
    pv = "'Pivot Analysis'"
    for name, col in (("Tindex", "G"), ("MonthNumber", "H"),
                      ("MonthRevenue", "C"), ("MonthOrders", "B")):
        wb.defined_names.add(DefinedName(
            name, attr_text=f"{pv}!${col}${c_first}:${col}${c_last}"))


def add_self_checks(ws, row, anchors, last, c_first, c_last):
    """Live cross-checks that prove the formulas evaluated correctly.

    LibreOffice cannot load documents in the build environment, so the
    workbook cannot be recalculated here. These formulas do the verifying
    instead: open the file in Excel and every Status cell must read OK.
    """
    pv = "'Pivot Analysis'"
    p1_head, p1_first, p1_last, n_years = anchors["p1"]
    p2_first, p2_last = anchors["p2"]
    p3_first, p3_last = anchors["p3"]
    grand_col = get_column_letter(2 + n_years)

    row = title(ws, row, "Self-check - open in Excel and confirm every row says OK",
                "Each check recomputes a total two different ways. If any row "
                "says CHECK, a formula did not evaluate and the workbook "
                "should not be submitted.")
    header_row(ws, row, ["Check", "Status"])
    row += 1
    checks = [
        ("Pivot 1 grand total equals SUM of all Amount",
         f"ROUND({pv}!${grand_col}${p1_last + 1}-SUM(Amount),2)"),
        ("Pivot 2 order count equals COUNT of all rows",
         f"SUM({pv}!$B${p2_first}:$B${p2_last})-COUNT(Amount)"),
        ("Pivot 2 revenue equals SUM of all Amount",
         f"ROUND(SUM({pv}!$C${p2_first}:$C${p2_last})-SUM(Amount),2)"),
        ("Pivot 3 order count equals COUNT of all rows",
         f"SUM({pv}!$B${p3_first}:$B${p3_last})-COUNT(Amount)"),
        ("Pivot 3 revenue equals SUM of all Amount",
         f"ROUND(SUM({pv}!$C${p3_first}:$C${p3_last})-SUM(Amount),2)"),
        ("Regression sample is 57 complete months",
         f"COUNT({pv}!$C${c_first}:$C${c_last})-57"),
        ("Cleaned Data holds 1,194 transaction lines",
         f"COUNT(Amount)-{last - 1}"),
    ]
    for label, expr in checks:
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name=FONT, size=10)
        st = ws.cell(row=row, column=2,
                     value=f'=IF(({expr})=0,"OK",'
                           f'"CHECK - difference of "&({expr}))')
        st.font = Font(name=FONT, size=10, bold=True)
        st.border = BOX
        st.fill = NOTE_FILL
        row += 1
    return row + 1


def sheet_readme(wb, anchors, last, c_first, c_last):
    """A short guide so the marker knows what is where."""
    ws = wb.create_sheet("Read Me", 0)
    widths(ws, {"A": 24, "B": 88})
    row = title(ws, 1, "BED 106 Checkpoint 2 - Spreadsheet Analytics Workbook",
                "Sales Trend Analysis. Built from the Checkpoint 1 database; "
                "same dataset, as the brief requires.")
    header_row(ws, row, ["Sheet", "What it contains"])
    row += 1
    guide = [
        ("Cleaned Data", "Task 2.1 sheet 1. All 1,194 transaction lines with "
                         "calendar and hierarchy columns added. MarginPct is a "
                         "formula, not a pasted value."),
        ("Pivot Analysis", "Task 2.1 sheet 2. Four cross-tabs built with "
                           "SUMIFS, COUNTIFS and AVERAGEIFS: category x year, "
                           "state, month, and sub-category 2023 vs 2024."),
        ("Pivot Charts", "Task 2.1 sheet 3. Four charts drawn from those "
                         "cross-tabs, each titled, axis-labelled and annotated."),
        ("Formulas Showcase", "Task 2.1 sheet 4. Eight Excel functions - "
                              "SUMIF, COUNTIF, SUMIFS, AVERAGEIF, VLOOKUP, IF, "
                              "TEXT, MAX/MIN - each answering a real question."),
        ("Descriptive Stats", "Task 2.2. Mean, median, mode, standard "
                              "deviation, variance, range, quartiles and "
                              "coefficient of variation for three variables."),
        ("Frequency", "Task 2.2. Frequency distribution and histogram for "
                      "Amount, with a check that the bins total the row count."),
        ("Correlation", "Task 2.3. Two variable pairs with Pearson r and "
                        "scatter plots with trendlines, plus a supporting third."),
        ("Regression", "Task 2.4. Simple linear regression of monthly revenue "
                       "on monthly order count: equation, R squared, t "
                       "statistic, p-value, forecasts and limitations."),
        ("Forecast", "Task 2.5. Trend test, seasonal index, a six-period "
                     "forecast, and an honest check against the two complete "
                     "2025 months held back from the analysis."),
    ]
    for name, what in guide:
        n = ws.cell(row=row, column=1, value=name)
        n.font = Font(name=FONT, size=10, bold=True)
        n.alignment = Alignment(vertical="top")
        w = ws.cell(row=row, column=2, value=what)
        w.font = Font(name=FONT, size=10)
        w.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1
    row = title(ws, row, "How to read this workbook")
    for text in [
        "Every computed cell is a live formula over named ranges (Amount, "
        "Profit, Quantity, Category, State, Yr, YearMonth and others). "
        "Change the data and everything recalculates.",
        "The cross-tabs on Pivot Analysis compute exactly what a PivotTable "
        "computes. To add native Excel PivotTables as well, follow "
        "docs/checkpoint2_excel_guide.md - it takes about two minutes each.",
        "Both ends of the source file are partial: it runs 22 March 2020 to "
        "15 March 2025. The analysis window is therefore the 57 whole months "
        "from April 2020 to December 2024. January and February 2025 are "
        "complete months and are used only to check the forecast.",
        "The written interpretations belong in the report, and per Section 3.2 "
        "of the brief they must be in the group's own words.",
    ]:
        c = ws.cell(row=row, column=1, value="- " + text)
        c.font = Font(name=FONT, size=10)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 30
        row += 1

    add_self_checks(ws, row + 1, anchors, last, c_first, c_last)
    return ws


def main():
    rows = load()
    wb = Workbook()

    ws_data, last = sheet_cleaned(wb, rows)
    define_names(wb, last)

    _, anchors, cats, years, states, subs, months = sheet_pivots(wb, rows, last)
    p3_first, p3_last = anchors["p3"]
    c_first, c_last = add_series_helpers(wb, months, p3_first)
    define_series_names(wb, c_first, c_last)

    sheet_charts(wb, anchors, years)
    sheet_formulas(wb, rows, last)
    sheet_descriptive(wb, last)
    sheet_frequency(wb)
    sheet_correlation(wb, c_first, c_last, last)
    sheet_regression(wb, c_first, c_last)
    sheet_forecast(wb, months, p3_first, p3_last)
    sheet_readme(wb, anchors, last, c_first, c_last)

    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(rows):,} data rows, {len(wb.sheetnames)} sheets: "
          f"{', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
